from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from database import PostgreSQLDatabase
from colorama import init, Fore, Style

# Инициализация colorama (необходимо вызывать один раз в начале программы)
init()
# Загрузка файла Excel
workbook = load_workbook(filename='ОН_ФЭВТ_1 курс.xlsx')

# Получение активного листа
sheet = workbook.active

db = PostgreSQLDatabase(host="localhost",
                        port="5432",
                        user='sergey',
                        password='Serez_Groza_1337',
                        database="vkr")


def move_work_zone_down(work_zone):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row += 3
    bottom_row += 3
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_right(work_zone):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_col += 4
    bottom_col += 4
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_up(work_zone):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row -= 18 * 6
    bottom_row -= 18 * 6
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_to_second_week(work_zone):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row = 116
    bottom_row = 118
    top_col = 8
    bottom_col = 11
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def get_current_group_name(work_zone) -> str:
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    top_row -= 1
    return str(sheet.cell(row=top_row, column=top_col).value).replace(" ", "").lower()


def check_cell_has_data(cell):
    if cell.value is not None:
        return True
    else:
        # Если ячейка объединена, проверяем, входит ли ее координата в список объединенных ячеек
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Получаем координаты первой ячейки объединенной области
                first_cell_row, first_cell_col = merged_range.min_row, merged_range.min_col
                first_cell = sheet.cell(row=first_cell_row, column=first_cell_col)
                if first_cell.value is not None:
                    return True
        return False


def move_cell_right(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на одну колонку вправо
    new_column = current_column + 1

    # Преобразуем номер колонки в буквенное представление
    new_column_letter = get_column_letter(new_column)

    # Формируем новый адрес ячейки
    new_cell_address = f"{new_column_letter}{current_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def move_cell_down(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на одну строку вниз (
    new_row = current_row + 3  # тут 3, потому что объединенные ячейки

    # Формируем новый адрес ячейки
    new_cell_address = f"{get_column_letter(current_column)}{new_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def move_cell_left_to_default(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на пять колонок влево
    new_column = current_column - 5

    # Преобразуем номер колонки в буквенное представление
    new_column_letter = get_column_letter(new_column)

    # Формируем новый адрес ячейки
    new_cell_address = f"{new_column_letter}{current_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def analyze_dates():
    db.connect()
    month_names = {
        "Январь": "01",
        "Февраль": "02",
        "Март": "03",
        "Апрель": "04",
        "Май": "05",
        "Июнь": "06",
        "Июль": "07",
        "Август": "08",
        "Сентябрь": "09",
        "Октябрь": "10",
        "Ноябрь": "11",
        "Декабрь": "12"
    }
    # названия месяцев
    month_dict = {}

    dates = {}

    work_cell = sheet['A6']
    # для записи в словарик названий месяцев
    for row in range(1, 5 + 1):
        month_dict.update({row: work_cell.value})
        work_cell = move_cell_right(work_cell)

    # ставим ячейку на начало дат
    work_cell = sheet['A7']
    # для всех недель
    for num_week in range(1, 2 + 1):
        # Цикл для всей недели
        for week_day in range(1, 7 + 1):
            # цикл для одного дня
            for column in range(1, 6 + 1):
                # для строки
                for row in range(1, 5 + 1):
                    # проверяем, что в строке число
                    if work_cell.value is not None and isinstance(work_cell.value, int) and work_cell.value > 0:
                        print(work_cell.value)
                        dates.update({month_dict[row]: work_cell.value})
                        current_month = month_names.get(month_dict[row])
                        date = f'2024-{current_month}-{work_cell.value}'
                        query = "INSERT INTO dates (date, week_day, week_num) VALUES (%s, %s, %s)"
                        db.execute_query(query, (date, week_day, num_week))
                    work_cell = move_cell_right(work_cell)
                work_cell = move_cell_left_to_default(work_cell)
                work_cell = move_cell_down(work_cell)
            print('----')
        work_cell = sheet["A116"]
    print(dates)
    db.disconnect()


def fill_lessons_table(week_num, week_day, group_name):
    # Выбираем даты из таблицы dates по номеру недели и дню недели
    query = "SELECT date FROM dates WHERE week_num = %s AND week_day = %s"
    dates = db.execute_query(query, (week_num, week_day))

    # Для каждой выбранной даты вставляем записи в таблицу lessons
    for date in dates:
        # Выполняем вставку записи для каждой группы
        insert_query = """
        INSERT INTO lessons (group_name, lesson_order, is_busy, lesson_date)
        VALUES (%s, %s, %s, %s)
        """
        # Здесь вы можете указать нужный порядковый номер пары, значение для is_busy и т. д.
        values = (group_name, 1, False, date[0])  # Предположим, что lesson_order = 1 и is_busy = False
        db.execute_query(insert_query, values)


def analyze_worksheet():
    db.connect()
    # Определение "рабочей зоны" (тут группа ячеек 4 на 3, в которой вся инфа для 1 пары 1 группы)
    work_zone = sheet['H7':'K9']

    # Словарь для групп (нужен для того, чтобы знать в какой колонке, какая группа)
    groups_dict = {}

    for num_week in range(1, 2 + 1):
        print(num_week)
        # Цикл для прохода по одной группе
        for number_group in range(10):
            # Название группы (Например прин-166)
            if groups_dict.get(number_group) is None:
                group_name = get_current_group_name(work_zone)
                groups_dict.update({number_group: group_name})
            else:
                group_name = groups_dict[number_group]
            print(Fore.GREEN + f'Название группы: {group_name}' + Style.RESET_ALL)
            # Добавить в бд название группы
            insert_query = """
                INSERT INTO groups (group_name, faculty, course)
                SELECT %s, %s, %s
                WHERE NOT EXISTS (
                    SELECT 1 FROM groups WHERE group_name = %s
                )
            """
            db.execute_query(insert_query, (group_name, "ФЭВТ", "1", group_name))

            # Цикл для прохода по всей неделе
            for week_day in range(1, 6 + 1):
                print(f"День недели: {week_day}")
                # Цикл для прохода по одному дню
                for number_para in range(1, 6 + 1):
                    print(f'Номер пары: {number_para}')
                    # Цикл для прохода по одной паре
                    has_lesson = False
                    for row in work_zone:
                        # Цикл для прохода по одной строке
                        if has_lesson:
                            break
                        for cell in row:
                            if check_cell_has_data(cell):
                                has_lesson = True
                                break
                        #     print(f'Значение в ячейке: {cell.value}')
                        # print("-------")

                    print(f'Есть ли занятие: {has_lesson}')
                    # Выбираем даты из таблицы dates по номеру недели и дню недели
                    query = "SELECT date FROM dates WHERE week_num = %s AND week_day = %s"
                    dates = db.execute_query(query, (num_week, week_day))
                    dates = dates[1:]
                    dates = dates[0]

                    # Для каждой выбранной даты вставляем записи в таблицу lessons
                    for date in dates:
                        insert_query = """
                            INSERT INTO lessons (group_name, lesson_order, is_busy, lesson_date)
                            VALUES (%s, %s, %s, %s)
                            """
                        # Здесь вы можете указать нужный порядковый номер пары, значение для is_busy и т. д.
                        values = (group_name, number_para, has_lesson, date[0])  # Предположим, что lesson_order = 1
                        # и is_busy = False
                        db.execute_query(insert_query, values)
                    print("_________________________")

                    work_zone = move_work_zone_down(work_zone)
            work_zone = move_work_zone_up(work_zone)
            work_zone = move_work_zone_right(work_zone)
        work_zone = move_work_zone_to_second_week(work_zone)
    db.disconnect()


if __name__ == '__main__':
    analyze_dates()
    analyze_worksheet()
