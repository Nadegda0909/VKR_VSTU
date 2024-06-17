from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from collections import defaultdict
from backend.database import PostgreSQLDatabase
import time

intervals_list = ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14', '15-16']


def fetch_lesson_intervals(db):
    sql_query = '''
    SELECT group_name, lesson_date, lesson_interval
    FROM lesson_intervals_for_ck
    ORDER BY group_name, lesson_date, lesson_interval;
    '''
    results = db.execute_query(sql_query)[1]

    lesson_intervals = defaultdict(list)
    for group_name, lesson_date, lesson_interval in results:
        lesson_intervals[group_name].append((lesson_date, lesson_interval))

    return lesson_intervals


def get_day_of_week(date):
    days = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ']
    return days[date.weekday()]


def get_time_slot(interval):
    slots = {
        '1-2': '8:30-10:00',
        '3-4': '10:10-11:40',
        '5-6': '11:50-13:20',
        '7-8': '13:40-15:10',
        '9-10': '15:20-16:50',
        '11-12': '17:00-18:30',
        '13-14': '18:35-20:00',
        '15-16': '20:05-21:30'
    }
    return slots[interval]


def map_interval(interval):
    mapping = {
        '1-2': ['1-2', '3-4'],
        '3-4': ['5-6', '7-8'],
        '5-6': ['9-10', '11-12'],
        '7-8': ['13-14', '15-16']
    }
    return mapping[interval]


def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    if not column_letter:
                        column_letter = cell.column_letter
                except Exception as e:
                    print(e)
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width


def make_schedule_excel(lesson_intervals):
    wb = Workbook()

    # Определяем стили
    header_font = Font(name='Times New Roman', size=12, bold=True)
    cell_font = Font(name='Times New Roman', size=12)
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    week_days = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ']
    time_slots = [get_time_slot(interval) for interval in intervals_list]

    for week_num in [1, 2]:
        ws = wb.create_sheet(title=f"Неделя {week_num}")

        # Заполняем заголовки
        for col_num, day in enumerate(['Начало пары'] + week_days, 1):
            cell = ws.cell(row=1, column=col_num, value=day)
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        for row_num, (interval, time_slot) in enumerate(zip(intervals_list, time_slots), 2):
            cell = ws.cell(row=row_num, column=1, value=f"{interval}\n({time_slot})")
            cell.font = header_font
            cell.border = border
            cell.alignment = wrap_alignment

        # Заполняем данными
        cell_content = defaultdict(lambda: defaultdict(list))  # Храним содержимое для каждой ячейки
        additional_columns = defaultdict(int)  # Счетчик дополнительных столбцов для каждого дня

        for group_name, intervals in lesson_intervals.items():
            for lesson_date, lesson_interval in intervals:
                week_query = '''
                SELECT week_num
                FROM learning_dates
                WHERE date = %s
                '''
                week = db.execute_query(week_query, (lesson_date,))[1][0][0]

                if week == week_num:
                    day = get_day_of_week(lesson_date)
                    day_col = week_days.index(day) + 2 + additional_columns[day]
                    excel_intervals = map_interval(lesson_interval)

                    for excel_interval in excel_intervals:
                        interval_row = intervals_list.index(excel_interval) + 2
                        date_str = lesson_date.strftime('%d.%m')
                        if ws.cell(row=interval_row, column=day_col).value:
                            additional_columns[day] += 1
                            day_col += 1
                        cell_content[(interval_row, day_col)][group_name].append(date_str)

        # Заполняем ячейки и делаем отдельные колонки для каждой группы
        max_additional_columns = max(additional_columns.values())
        for (interval_row, day_col), groups in cell_content.items():
            cell_text = "\n".join([f"{group_name} ({', '.join(dates)})" for group_name, dates in groups.items()])
            cell = ws.cell(row=interval_row, column=day_col, value=cell_text)
            cell.font = cell_font
            cell.border = border
            cell.alignment = wrap_alignment

        # Растягиваем заголовки дней недели
        for day_index in range(len(week_days)):
            day_col = day_index + 2
            for offset in range(max_additional_columns + 1):
                cell = ws.cell(row=1, column=day_col + offset, value=week_days[day_index])
                cell.font = header_font
                cell.border = border
                cell.alignment = center_alignment

        # Автовыравнивание ширины столбцов
        adjust_column_width(ws)

    # Удаляем стандартный лист, созданный по умолчанию
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    wb.save("schedule_ck.xlsx")


if __name__ == '__main__':
    t = time.time()
    db = PostgreSQLDatabase()
    db.connect()

    lesson_intervals = fetch_lesson_intervals(db)
    make_schedule_excel(lesson_intervals)

    db.disconnect()
    print("--- %s seconds --- schedule_maker" % (time.time() - t))
