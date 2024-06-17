from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from collections import defaultdict
from backend.database import PostgreSQLDatabase
import time


def select_ck_group_university_group(db):
    sql_query = '''
    SELECT sc.ck_group, sc.oop_group_2023_2024, sc.university_name, sc.faculty
    FROM students_ck sc 
    ORDER BY sc.ck_group, sc.oop_group_2023_2024
    '''

    # Выполнение SQL-запроса и получение результатов
    results = db.execute_query(sql_query)[1]

    # Инициализация пустого словаря
    ck_group_university_group = defaultdict(list)
    group_university_info = {}

    # Заполняем словарь данными из запроса
    for ck_group, oop_group, university_name, faculty in results:
        if oop_group not in ck_group_university_group[ck_group]:
            ck_group_university_group[ck_group].append(oop_group)
        if university_name.lower() != 'волггту' or 'иаис' in university_name.lower() or 'иаис' in faculty.lower():
            group_university_info[oop_group] = university_name

    return ck_group_university_group, group_university_info


def get_program_from_ck_group(ck_group):
    # Функция для получения названия программы из имени группы ЦК
    return ck_group.split('_')[0]


def adjust_column_width(ws):
    # Автоматически корректируем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    if not column_letter:
                        column_letter = cell.column_letter
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width


def make_excel(ck_group_university_group, group_university_info):
    # Создаем новую книгу
    wb = Workbook()

    # Определяем стили
    header_font = Font(name='Times New Roman', size=12, bold=True)
    cell_font = Font(name='Times New Roman', size=12)
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Группируем ck_group по направлениям
    program_ck_groups = defaultdict(list)
    for ck_group in ck_group_university_group:
        program = get_program_from_ck_group(ck_group)
        program_ck_groups[program].append(ck_group)

    # Создаем отдельный лист для каждого направления
    for program, ck_groups in program_ck_groups.items():
        ws = wb.create_sheet(title=program)

        # Найти максимальное количество групп в группе ЦК
        max_groups = max(len(ck_group_university_group[ck_group]) for ck_group in ck_groups)

        # Заполняем заголовки
        headers = ['Группа ЦК', 'Группы', 'Число студентов']
        ws.cell(row=1, column=1, value=headers[0]).font = header_font
        ws.cell(row=1, column=1).border = border
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + max_groups)
        header_cell = ws.cell(row=1, column=2, value=headers[1])
        header_cell.font = header_font
        header_cell.border = border
        header_cell.alignment = center_alignment
        ws.cell(row=1, column=2 + max_groups, value=headers[2]).font = header_font
        ws.cell(row=1, column=2 + max_groups).border = border

        # Заполняем данными
        row = 2
        for ck_group in ck_groups:
            ck_cell = ws.cell(row=row, column=1, value=ck_group)
            ck_cell.font = header_font
            ck_cell.fill = orange_fill
            ck_cell.border = border

            for col, university_group in enumerate(ck_group_university_group[ck_group], start=2):
                if university_group in group_university_info:
                    cell_value = f"{university_group} ({group_university_info[university_group]})"
                else:
                    cell_value = university_group
                cell = ws.cell(row=row, column=col, value=cell_value)
                cell.font = cell_font
                cell.border = border

            # Считаем количество студентов в группе ЦК
            student_count_query = '''
            SELECT COUNT(*)
            FROM students_ck
            WHERE ck_group = %s
            '''
            student_count = db.execute_query(student_count_query, (ck_group,))[1][0][0]

            student_count_cell = ws.cell(row=row, column=2 + max_groups, value=student_count)
            student_count_cell.font = cell_font
            student_count_cell.border = border

            row += 1

        # Автовыравнивание ширины столбцов для текущего листа
        adjust_column_width(ws)

    # Устанавливаем шрифт и рамки для всех ячеек
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = cell_font
                cell.border = border

    # Удаляем стандартный лист, созданный по умолчанию
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    # Сохраняем книгу
    wb.save("group_ck.xlsx")


if __name__ == '__main__':
    t = time.time()
    db = PostgreSQLDatabase()
    db.connect()

    ck_group_university_group_dict, group_university_info = select_ck_group_university_group(db)
    make_excel(ck_group_university_group_dict, group_university_info)

    db.disconnect()
    print("--- %s seconds --- group_ck_maker" % (time.time() - t))
