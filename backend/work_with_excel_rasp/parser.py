import os
import re
import shutil
import time
from datetime import datetime
import itertools

from concurrent.futures import ProcessPoolExecutor

from colorama import init, Fore, Style
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from collections import defaultdict
from backend.database import PostgreSQLDatabase
from backend.work_with_excel_rasp.downloader import download_schedule_files, convert_xls_to_xlsx



def find_cells(sheet, values_to_find):
    # Создаем пустой список для хранения найденных ячеек с их координатами и номерами колонок
    found_cells = []

    # Проходимся по всем ячейкам и ищем нужные значения
    for row in sheet.iter_rows():
        for col_idx, cell in enumerate(row, start=1):
            cell_value = str(cell.value).replace(" ", "") if cell.value else ""

            if len(values_to_find) == 1:
                # Удаляем пробелы и проверяем наличие значения
                if cell_value == values_to_find[0]:
                    if col_idx == 7:
                        found_cells.append((cell.coordinate, col_idx))
            else:
                first_value, second_value = values_to_find
                if cell_value == first_value or cell_value == second_value:
                    found_cells.append((cell.coordinate, col_idx))

    # Возвращаем список найденных ячеек
    return found_cells


def create_work_zone(sheet, coordinate):
    # Получаем координаты начальной ячейки
    start_row, start_col = coordinate_to_tuple(coordinate)

    start_col += 1

    # Смещаемся на 4 ячейки вправо
    end_col = start_col + 3

    # Создаем адреса начальной и конечной ячеек для рабочей зоны
    start_cell = f"{get_column_letter(start_col)}{start_row}"
    end_cell = f"{get_column_letter(end_col)}{start_row + 2}"

    # Получаем рабочую зону
    work_zone = sheet[start_cell:end_cell]

    return work_zone


def move_work_zone_down(work_zone, sheet):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row += 3
    bottom_row += 3
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_right(work_zone, sheet):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_col += 4
    bottom_col += 4
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_up(work_zone, sheet):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row -= 18 * 6
    bottom_row -= 18 * 6
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def get_current_group_name(work_zone, sheet) -> str:
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    top_row -= 1
    # Получаем название группы в формате "прин - 266", то есть в "сыром" виде
    group_name = str(sheet.cell(row=top_row, column=top_col).value).strip().lower()
    if not (group_name == 'none' or group_name == '' or group_name == ' '):
        # Если в названии группы есть "-", то пробуем поделить через "-"
        group_name = group_name.split('-')  # Тут получается список в формате ["прин ", " 266"]
        if len(group_name) > 1:
            for i in range(len(group_name)):
                group_name[i] = group_name[i].replace('дистанционно', '')
                group_name[i] = group_name[i].strip()
            group_name = group_name[0] + '-' + group_name[1]
            return group_name  # Возвращаем в формате "прин-266"

        # Если "-" не оказалось, то пробуем поделить через пробел
        group_name = group_name[0].split(' ')
        for i in range(len(group_name)):
            group_name[i] = group_name[i].strip()
        # Для случаев, когда несколько групп пишут через запятую
        if ',' in group_name[1]:
            # в этом случае возвращаем название групп в формате "прин-266 прин-277"
            return group_name[0] + '-' + group_name[1][:-1] + ' ' + group_name[2] + '-' + group_name[3]
        group_name = group_name[0] + '-' + group_name[1]
        return group_name  # Возвращаем в формате "прин-266"
    return 'none'


def check_dates(input_string: str) -> list | None:
    if isinstance(input_string, str):
        current_year = datetime.now().year
        input_string = input_string.lower()
        date_pattern = r"\d{2}\.\d{2}"
        dates_strings = re.findall(date_pattern, input_string)
        if len(dates_strings) == 0:
            return None
        dates = [datetime.strptime(date + f".{current_year}", "%d.%m.%Y").date() for date in dates_strings]
        return dates
    return None


def check_count_pars(value):
    if isinstance(value, str):
        # Паттерн для поиска числа, за которым следует слово "час" или "часа"
        pattern = r"(\d+)\s+(часа|час)"

        # Поиск шаблона в значении ячейки
        match = re.search(pattern, value)

        # Если найден шаблон
        if match:
            # Извлекаем число из найденного шаблона
            count = int(match.group(1))

            # Делим число на 2 и возвращаем результат
            return count // 2

    # Если не найден шаблон, возвращаем None
    return None


def check_cell_has_data_and_dates(cell, sheet):
    if cell.value is not None:
        dates = check_dates(cell.value)
        count_pars = check_count_pars(cell.value)
        return True, dates, count_pars
    else:
        # Если ячейка объединена, проверяем, входит ли ее координата в список объединенных ячеек
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Получаем координаты первой ячейки объединенной области
                first_cell_row, first_cell_col = merged_range.min_row, merged_range.min_col
                first_cell = sheet.cell(row=first_cell_row, column=first_cell_col)
                if first_cell.value is not None:
                    dates = check_dates(first_cell.value)
                    count_pars = check_count_pars(first_cell.value)
                    return True, dates, count_pars
        return False, None, None


def move_cell_right(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на одну колонку вправо
    new_column = current_column + 1

    # Преобразуем номер колонки в буквенное представление
    new_column_letter = get_column_letter(new_column)

    # Формируем новый адрес ячейки
    new_cell_address = f"{new_column_letter}{current_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def move_cell_down(cell, count):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на одну строку вниз (
    new_row = current_row + count

    # Формируем новый адрес ячейки
    new_cell_address = f"{get_column_letter(current_column)}{new_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def move_cell_to_leftmost(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row

    # Получаем буквенное представление крайней левой колонки
    leftmost_column_letter = get_column_letter(1)

    # Формируем новый адрес ячейки крайней левой колонки
    new_cell_address = f"{leftmost_column_letter}{current_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def analyze_dates(
        filename='./converted_files/Бакалавриат, специалитет/Факультет автоматизированных систем, транспорта и вооружений/ОН_ФАСТИВ_3 курс (320-324).xlsx'):  # его берем за эталон

    # Загрузка файла Excel
    workbook = load_workbook(filename=filename)

    # Получение активного листа
    sheet = workbook.active
    month_names = {
        "январь": "01",
        "февраль": "02",
        "март": "03",
        "апрель": "04",
        "май": "05",
        "июнь": "06",
        "июль": "07",
        "август": "08",
        "сентябрь": "09",
        "октябрь": "10",
        "ноябрь": "11",
        "декабрь": "12"
    }
    # названия месяцев
    month_dict = {}

    dates = {}
    cell_coordinates = find_cells(sheet, ("февраль", "сентябрь"))
    cell_coordinate = cell_coordinates[0][0]
    work_cell = sheet[cell_coordinate]
    # для записи в словарик названий месяцев
    for row in range(1, 5 + 1):
        month_dict.update({row: work_cell.value.lower()})
        work_cell = move_cell_right(work_cell)

    # ставим ячейку на начало дат
    work_cell = move_cell_to_leftmost(work_cell)
    work_cell = move_cell_down(work_cell, 1)
    # для всех недель
    for num_week in range(1, 2 + 1):
        # Цикл для всей недели
        for week_day in range(1, 7 + 1):
            # цикл для одного дня
            for column in range(1, 6 + 1):
                # для строки
                for row in range(1, 5 + 1):
                    # проверяем, что в строке число
                    if work_cell.value is not None and isinstance(work_cell.value, int) and work_cell.value > 0:
                        # print(work_cell.value)
                        dates.update({month_dict[row]: work_cell.value})
                        current_month = month_names.get(month_dict[row])
                        date = f'2024-{current_month}-{work_cell.value}'
                        query = """
                            INSERT INTO learning_dates (date, week_day, week_num)
                            VALUES (%s, %s, %s)
                            -- ON CONFLICT (date) DO NOTHING
                        """

                        db.execute_query(query, (date, week_day, num_week))

                    work_cell = move_cell_right(work_cell)
                work_cell = move_cell_to_leftmost(work_cell)
                if week_day == 7:
                    continue
                work_cell = move_cell_down(work_cell, 3)
            # print('----')
        work_cell = move_cell_down(work_cell, 1)
    # print(dates)
    print(f"{Fore.GREEN}Даты сформированы! {Style.RESET_ALL}")


def get_course(group_name):
    # Разделяем строку по символу "-"
    parts = group_name.split("-")

    # Берем последний элемент списка, то есть часть с номером курса,
    # и берем его первый символ, который должен быть номером курса
    course = parts[-1][0]

    return course


def get_faculty(filename):
    # Разделяем имя файла по символу "_" и берем вторую часть,
    # которая содержит название факультета
    faculty = filename.split("_")[1]
    if "Магистратура" in faculty:
        faculty = filename.split('_')[-1].split(' ')[-1].replace('.xlsx', '')

    return faculty


def get_study_program(filepath):
    if "бакалавриат" in filepath.lower():
        return "бакалавриат/специалитет"
    return "магистратура"


def insert_schedule_for_one_day_in_db(schedule, num_week, week_day, group_name):
    db = PostgreSQLDatabase()
    db.connect()
    for number_para in range(1, 6 + 1):
        lesson_dates = schedule[f'lesson_{number_para}']["dates"]
        has_lesson = schedule[f'lesson_{number_para}']["has_lesson"]
        # Выбираем даты из таблицы dates по номеру недели и дню недели
        query = "SELECT date FROM learning_dates WHERE week_num = %s AND week_day = %s"
        dates_from_db = db.execute_query(query, (num_week, week_day))
        dates_from_db = dates_from_db[1:][0]
        dates_from_db = [date[0] for date in dates_from_db]  # Преобразуем список кортежей в список значений
        if not lesson_dates:  # Если у пары нет дат, то берем данные из бд
            dates = dates_from_db
            # Для каждой выбранной даты вставляем записи в таблицу lessons
            for date in dates:
                insert_query = """
                            INSERT INTO lessons_for_vstu (group_name, lesson_order, is_busy, lesson_date)
                            SELECT %s, %s, %s, %s
                            WHERE NOT EXISTS (
                                SELECT 1 FROM lessons_for_vstu 
                                WHERE group_name = %s AND lesson_order = %s AND lesson_date = %s
                            )
                        """
                values = (group_name, number_para, has_lesson, date, group_name, number_para, date)
                db.execute_query(insert_query, values)
        else:
            dates = list(
                set(lesson_dates) & set(dates_from_db))  # Если есть даты, то берем пересечение этих дат, с датами из бд

            # Для каждой выбранной даты вставляем записи в таблицу lessons
            for date in dates:
                insert_query = """
                    INSERT INTO lessons_for_vstu (group_name, lesson_order, is_busy, lesson_date)
                    SELECT %s, %s, %s, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM lessons_for_vstu 
                        WHERE group_name = %s AND lesson_order = %s AND lesson_date = %s
                    )
                """
                values = (group_name, number_para, has_lesson, date, group_name, number_para, date)
                db.execute_query(insert_query, values)

            dates = list(set(dates_from_db) - (set(lesson_dates) & set(dates_from_db)))
            for date in dates:
                insert_query = """
                    INSERT INTO lessons_for_vstu (group_name, lesson_order, is_busy, lesson_date)
                    SELECT %s, %s, %s, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM lessons_for_vstu 
                        WHERE group_name = %s AND lesson_order = %s AND lesson_date = %s
                    )
                """
                values = (group_name, number_para, not has_lesson, date, group_name, number_para, date)
                db.execute_query(insert_query, values)


def update_dates(lesson, new_dates):
    lesson['dates'].extend(new_dates)
    lesson['dates'] = list(set(lesson['dates']))


def analyze_excel_file(filepath, filename):
    db = PostgreSQLDatabase()
    db.connect()
    # Загрузка файла Excel
    workbook = load_workbook(filename=filepath)  # Сомнительно, но окей
    # Получить тип учебной программы (Бакалавриат/Специалитет или Магистратура)
    study_program = get_study_program(filepath)
    # Узнать факультет
    faculty = get_faculty(filename)

    if 'магистратура' in study_program:  # Временная штука для отсечения магистрантов
        return

    # Т.К. курс вычисляется из номера группы, а номера группы у нас на данный момент нет, то ставим ему None
    course = None

    # Получение активного листа
    sheet = workbook.active
    # Определение "рабочей зоны" (тут группа ячеек 4 на 3, в которой вся инфа для 1 пары 1 группы)
    cell_coordinates = find_cells(sheet, ("1-2",))

    coordinate = cell_coordinates[0][0]
    work_zone = create_work_zone(sheet, coordinate)

    # Словарь для групп (нужен для того, чтобы знать в какой колонке, какая группа)
    groups_dict = {}
    print(f'{Fore.BLUE}Факультет: {faculty}{Style.RESET_ALL}')
    for num_week in range(1, 2 + 1):
        # print(num_week)
        # Цикл для прохода по одной группе
        for number_group in range(1, 10):
            # Название группы (Например прин-166)
            if groups_dict.get(number_group) is None:
                group_name = get_current_group_name(work_zone, sheet)
                groups_dict.update({number_group: group_name})
            else:
                group_name = groups_dict[number_group]
            if group_name == 'none' or group_name == '' or group_name == ' ':
                continue
            group_name_list = group_name.split(' ')
            # Если в ячейке написаны 2 группы, то прогоняем цикл по одному и тому же расписанию 2 раза, но с разными
            # названиями групп
            for group_index in range(len(group_name_list)):
                group_name = group_name_list[group_index]
                print(Fore.GREEN + f'Название группы: {group_name}' + Style.RESET_ALL)
                # if group_name != 'мап-450':
                #     continue
                if course is None:
                    course = get_course(group_name)
                # Добавить в бд название группы
                insert_query = """
                    INSERT INTO groups_vstu_and_others (group_name, faculty, course, program)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (group_name) DO NOTHING
                """
                db.execute_query(insert_query, (group_name, faculty, course, study_program))

                # Цикл для прохода по всей неделе
                for week_day in range(1, 6 + 1):
                    # print(f"День недели: {week_day}")
                    # Цикл для прохода по одному дню
                    schedule_for_one_day = {
                        'lesson_1': {
                            "duration": 1,
                            "dates": [],
                            "has_lesson": False
                        },
                        'lesson_2': {
                            "duration": 1,
                            "dates": [],
                            "has_lesson": False
                        },
                        'lesson_3': {
                            "duration": 1,
                            "dates": [],
                            "has_lesson": False
                        },
                        'lesson_4': {
                            "duration": 1,
                            "dates": [],
                            "has_lesson": False
                        },
                        'lesson_5': {
                            "duration": 1,
                            "dates": [],
                            "has_lesson": False
                        },
                        'lesson_6': {
                            "duration": 1,
                            "dates": [],
                            "has_lesson": False
                        },
                    }
                    for number_para in range(1, 6 + 1):
                        # print(f'Номер пары: {number_para}')
                        new_has_lesson = False
                        new_lesson_dates = []
                        new_count_pars = None
                        # Цикл для прохода по одной паре
                        for row in work_zone:
                            # Цикл для прохода по одной строке
                            for cell in row:
                                has_lesson, lesson_dates, count_pars = check_cell_has_data_and_dates(cell, sheet)
                                if has_lesson:
                                    new_has_lesson = True
                                if lesson_dates:
                                    new_lesson_dates.extend(lesson_dates)
                                if count_pars is not None:
                                    new_count_pars = count_pars
                        if new_has_lesson is True:
                            schedule_for_one_day[f'lesson_{number_para}']['has_lesson'] = new_has_lesson

                        if new_lesson_dates:
                            if number_para > 1:
                                prev_lesson = schedule_for_one_day[f'lesson_{number_para - 1}']
                                current_lesson = schedule_for_one_day[f'lesson_{number_para}']

                                if prev_lesson['duration'] > 1:
                                    update_dates(prev_lesson, new_lesson_dates)
                                    current_lesson['dates'] = list(set(prev_lesson['dates']))
                                else:
                                    update_dates(current_lesson, new_lesson_dates)
                            else:
                                current_lesson = schedule_for_one_day[f'lesson_{number_para}']
                                update_dates(current_lesson, new_lesson_dates)

                        if new_count_pars is not None:
                            schedule_for_one_day[f'lesson_{number_para}']['duration'] = new_count_pars

                        work_zone = move_work_zone_down(work_zone, sheet)
                    # print(schedule_for_one_day)
                    insert_schedule_for_one_day_in_db(schedule_for_one_day, num_week, week_day, group_name)
                work_zone = move_work_zone_up(work_zone, sheet)
            work_zone = move_work_zone_right(work_zone, sheet)
        coordinate = cell_coordinates[6][0]
        work_zone = create_work_zone(sheet, coordinate)
        # print(work_zone)


def process_file(file_path, file):
    # print(f"Processing file: {file_path}")
    analyze_excel_file(file_path, file)


def analyze_excel_files_in_folder(folder_path):
    exclude_prefix = '~$'  # Префикс для исключения файлов
    # Используем стек для хранения путей к папкам
    folders_path_stack = [folder_path]
    files_to_process = []

    while folders_path_stack:
        current_folder_path = folders_path_stack.pop()
        if 'инженерных кадров' in current_folder_path:
            continue
        # Получаем список файлов и папок в текущей папке
        files = os.listdir(current_folder_path)
        for file in files:
            file_path = os.path.join(current_folder_path, file)
            if os.path.isdir(file_path):  # Если это папка
                folders_path_stack.append(file_path)
            elif file.endswith('.xlsx') and not file.startswith(exclude_prefix) and not file.startswith('ОС_'):
                # Обработка только файлов с расширением .xlsx и не начинающихся с указанного префикса
                files_to_process.append((file_path, file))

    # Обработка файлов с использованием ProcessPoolExecutor
    with ProcessPoolExecutor() as executor:
        futures = [executor.submit(process_file, file_path, file) for file_path, file in files_to_process]

        for future in futures:
            future.result()
            # try:
            #       # Вызывает исключения, если они произошли
            # except Exception as e:
            #     print(f"Error processing file: {e}")

    print(f"{Fore.GREEN}В базу занесены все группы!{Style.RESET_ALL}")


def delete_files_and_download_files():
    items = os.listdir()
    folders = [item for item in items if os.path.isdir(item) and "__" not in item]
    for folder in folders:
        try:
            shutil.rmtree(folder)
        except OSError as e:
            print(f'{Fore.RED}Ошибка при удалении папки {folder} {e} {Style.RESET_ALL}')
    download_schedule_files()
    convert_xls_to_xlsx()


def create_table_lesson_intervals():
    print(f'{Fore.BLUE}Создается таблица с интервалами групп{Style.RESET_ALL}')
    db = PostgreSQLDatabase()
    db.connect()
    insert_query = '''
    select distinct l.group_name, lesson_order, is_busy, d.week_day, week_num, date
    from lessons_for_vstu l
    left join public.learning_dates d on d.date = l.lesson_date
    order by group_name,week_num, week_day, date, lesson_order
    '''
    lessons = db.execute_query(insert_query)[1]
    for i in range(1, len(lessons)):
        current_lesson = lessons[i]
        current_lesson_number_para = current_lesson[1]
        current_lesson_is_busy = current_lesson[2]
        current_lesson_group_name = current_lesson[0]

        previous_lesson = lessons[i - 1]
        previous_lesson_group_name = previous_lesson[0]
        previous_lesson_number_para = previous_lesson[1]
        previous_lesson_is_busy = previous_lesson[2]

        if (previous_lesson_number_para == 6) or (previous_lesson_group_name != current_lesson_group_name):
            continue

        is_busy = current_lesson_is_busy or previous_lesson_is_busy
        lesson_interval = f'{previous_lesson_number_para}-{current_lesson_number_para}'
        # print(current_lesson)
        insert_query = '''
        insert into lesson_intervals_for_vstu (group_name, lesson_interval, lesson_date, is_busy) VALUES (%s, %s, %s, %s)
        '''
        db.execute_query(insert_query, (previous_lesson_group_name, lesson_interval, previous_lesson[-1], is_busy))


if __name__ == '__main__':
    t = time.time()
    # Инициализация colorama (необходимо вызывать один раз в начале программы)
    init()

    db = PostgreSQLDatabase()
    delete_files_and_download_files()
    db.connect()
    db.truncate_table('lessons_for_vstu')
    db.truncate_table('groups_vstu_and_others')
    db.truncate_table('learning_dates')
    db.truncate_table('lesson_intervals_for_vstu')
    analyze_dates()
    analyze_excel_files_in_folder(
        'converted_files/')
    create_table_lesson_intervals()
    print("--- %s seconds --- parser" % (time.time() - t))
