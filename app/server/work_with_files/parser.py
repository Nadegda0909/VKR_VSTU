import os
import re
import shutil
import time
from datetime import datetime

from colorama import init, Fore, Style
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

from app.server.database import PostgreSQLDatabase
from app.server.work_with_files.downloader import download_schedule_files, convert_xls_to_xlsx


def find_cell(sheet, value_to_find, position):
    # Проходимся по всем ячейкам и ищем нужное значение
    for row in sheet.iter_rows():
        for cell in row:
            if len(value_to_find) == 1:
                # Удаляем пробелы и проверяем наличие значения
                if cell.value and str(cell.value).replace(" ", "") == value_to_find[0]:
                    if position == 2:
                        continue
                    return cell.coordinate
            else:
                first_date, second_date = value_to_find
                if cell.value and (str(cell.value).replace(" ", "") == first_date or
                                   str(cell.value).replace(" ", "") == second_date):
                    if position == 2:
                        continue
                    return cell.coordinate

    # Если значение не найдено, возвращаем None
    return None


def create_work_zone(sheet, occurrence_number):
    # Получаем координаты начальной ячейки
    start_row, start_col = coordinate_to_tuple(find_cell(sheet, ("1-2",), occurrence_number))

    start_col += 1

    # Смещаемся на 4 ячейки вправо
    end_col = start_col + 3

    # Создаем адреса начальной и конечной ячеек для рабочей зоны
    start_cell = f"{get_column_letter(start_col)}{start_row}"
    end_cell = f"{get_column_letter(end_col)}{start_row + 2}"

    # Получаем рабочую зону
    work_zone = sheet[start_cell:end_cell]

    return work_zone


def move_work_zone_down(work_zone, sheet):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row += 3
    bottom_row += 3
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_right(work_zone, sheet):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_col += 4
    bottom_col += 4
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def move_work_zone_up(work_zone, sheet):
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    bottom_row, bottom_col = coordinate_to_tuple(work_zone[-1][-1].coordinate)
    top_row -= 18 * 6
    bottom_row -= 18 * 6
    top_cell = f"{get_column_letter(top_col)}{top_row}"
    bottom_cell = f"{get_column_letter(bottom_col)}{bottom_row}"
    work_zone = sheet[top_cell:bottom_cell]
    return work_zone


def get_current_group_name(work_zone, sheet) -> str:
    top_row, top_col = coordinate_to_tuple(work_zone[0][0].coordinate)
    top_row -= 1
    # Получаем название группы в формате "прин - 266", то есть в "сыром" виде
    group_name = str(sheet.cell(row=top_row, column=top_col).value).strip().lower()
    if group_name != 'none':
        # Если в названии группы есть "-", то пробуем поделить через "-"
        group_name = group_name.split('-')  # Тут получается список в формате ["прин ", " 266"]
        if len(group_name) > 1:
            for i in range(len(group_name)):
                group_name[i] = group_name[i].strip()
            group_name = group_name[0] + '-' + group_name[1]
            return group_name  # Возвращаем в формате "прин-266"

        # Если "-" не оказалось, то пробуем поделить через пробел
        group_name = group_name[0].split(' ')
        for i in range(len(group_name)):
            group_name[i] = group_name[i].strip()
        # Для случаев, когда несколько групп пишут через запятую
        if ',' in group_name[1]:
            # в этом случае возвращаем название групп в формате "прин-266 прин-277"
            return group_name[0] + '-' + group_name[1][:-1] + ' ' + group_name[2] + '-' + group_name[3]
        group_name = group_name[0] + '-' + group_name[1]
        return group_name  # Возвращаем в формате "прин-266"
    return 'none'


def check_dates(input_string: str) -> list | None:
    if isinstance(input_string, str):
        current_year = datetime.now().year
        input_string = input_string.lower()
        date_pattern = r"\d{2}\.\d{2}"
        dates_strings = re.findall(date_pattern, input_string)
        if len(dates_strings) == 0:
            return None
        dates = [datetime.strptime(date + f".{current_year}", "%d.%m.%Y") for date in dates_strings]
        return dates
    return None


def check_count_pars(value):
    if isinstance(value, str):
        # Паттерн для поиска числа, за которым следует слово "час" или "часа"
        pattern = r"(\d+)\s+(часа|час)"

        # Поиск шаблона в значении ячейки
        match = re.search(pattern, value)

        # Если найден шаблон
        if match:
            # Извлекаем число из найденного шаблона
            count = int(match.group(1))

            # Делим число на 2 и возвращаем результат
            return count / 2

    # Если не найден шаблон, возвращаем None
    return None


def check_cell_has_data_and_dates(cell, sheet):
    if cell.value is not None:
        dates = check_dates(cell.value)
        count_pars = check_count_pars(cell.value)
        return True, dates, count_pars
    else:
        # Если ячейка объединена, проверяем, входит ли ее координата в список объединенных ячеек
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Получаем координаты первой ячейки объединенной области
                first_cell_row, first_cell_col = merged_range.min_row, merged_range.min_col
                first_cell = sheet.cell(row=first_cell_row, column=first_cell_col)
                if first_cell.value is not None:
                    dates = check_dates(first_cell.value)
                    count_pars = check_count_pars(first_cell.value)
                    return True, dates, count_pars
        return False, None, None


def move_cell_right(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на одну колонку вправо
    new_column = current_column + 1

    # Преобразуем номер колонки в буквенное представление
    new_column_letter = get_column_letter(new_column)

    # Формируем новый адрес ячейки
    new_cell_address = f"{new_column_letter}{current_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def move_cell_down(cell, count):
    # Получаем координаты текущей ячейки
    current_row = cell.row
    current_column = cell.column

    # Смещаемся на одну строку вниз (
    new_row = current_row + count

    # Формируем новый адрес ячейки
    new_cell_address = f"{get_column_letter(current_column)}{new_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def move_cell_to_leftmost(cell):
    # Получаем координаты текущей ячейки
    current_row = cell.row

    # Получаем буквенное представление крайней левой колонки
    leftmost_column_letter = get_column_letter(1)

    # Формируем новый адрес ячейки крайней левой колонки
    new_cell_address = f"{leftmost_column_letter}{current_row}"

    # Получаем ячейку по новому адресу
    new_cell = cell.parent[new_cell_address]

    return new_cell


def analyze_dates(filename='./converted_files/Бакалавриат, специалитет/Факультет электроники и вычислительной '
                           'техники/ОН_ФЭВТ_2 курс.xlsx'):  # его берем за эталон
    # Загрузка файла Excel
    workbook = load_workbook(filename=filename)

    # Получение активного листа
    sheet = workbook.active
    db.connect()
    month_names = {
        "январь": "01",
        "февраль": "02",
        "март": "03",
        "апрель": "04",
        "май": "05",
        "июнь": "06",
        "июль": "07",
        "август": "08",
        "сентябрь": "09",
        "октябрь": "10",
        "ноябрь": "11",
        "декабрь": "12"
    }
    # названия месяцев
    month_dict = {}

    dates = {}

    work_cell = sheet[find_cell(sheet, ("февраль", "сентябрь"), 1)]
    # для записи в словарик названий месяцев
    for row in range(1, 5 + 1):
        month_dict.update({row: work_cell.value.lower()})
        work_cell = move_cell_right(work_cell)

    # ставим ячейку на начало дат
    work_cell = move_cell_to_leftmost(work_cell)
    work_cell = move_cell_down(work_cell, 1)
    # для всех недель
    for num_week in range(1, 2 + 1):
        # Цикл для всей недели
        for week_day in range(1, 7 + 1):
            # цикл для одного дня
            for column in range(1, 6 + 1):
                # для строки
                for row in range(1, 5 + 1):
                    # проверяем, что в строке число
                    if work_cell.value is not None and isinstance(work_cell.value, int) and work_cell.value > 0:
                        # print(work_cell.value)
                        dates.update({month_dict[row]: work_cell.value})
                        current_month = month_names.get(month_dict[row])
                        date = f'2024-{current_month}-{work_cell.value}'
                        query = """
                            INSERT INTO dates (date, week_day, week_num)
                            VALUES (%s, %s, %s)
                            -- ON CONFLICT (date) DO NOTHING
                        """
                        db.execute_query(query, (date, week_day, num_week))

                    work_cell = move_cell_right(work_cell)
                work_cell = move_cell_to_leftmost(work_cell)
                if week_day == 7:
                    continue
                work_cell = move_cell_down(work_cell, 3)
            # print('----')
        work_cell = move_cell_down(work_cell, 1)
    # print(dates)
    print(f"{Fore.GREEN}Даты сформированы! {Style.RESET_ALL}")
    db.disconnect()


def get_course(group_name):
    # Разделяем строку по символу "-"
    parts = group_name.split("-")

    # Берем последний элемент списка, то есть часть с номером курса,
    # и берем его первый символ, который должен быть номером курса
    course = parts[-1][0]

    return course


def get_faculty(filename):
    # Разделяем имя файла по символу "_" и берем вторую часть,
    # которая содержит название факультета
    faculty = filename.split("_")[1]
    if "Магистратура" in faculty:
        faculty = filename.split('_')[-1].split(' ')[-1].replace('.xlsx', '')

    return faculty


def get_study_program(filepath):
    if "бакалавриат" in filepath.lower():
        return "бакалавриат/специалитет"
    return "магистратура"


def analyze_excel_file(filepath, filename):
    # Загрузка файла Excel
    workbook = load_workbook(filename=filepath)  # Сомнительно, но окей
    # Получить тип учебной программы (Бакалавриат/Специалитет или Магистратура)
    study_program = get_study_program(filepath)
    # Узнать факультет
    faculty = get_faculty(filename)

    # Т.К. курс вычисляется из номера группы, а номера группы у нас на данный момент нет, то ставим ему None
    course = None

    # Получение активного листа
    sheet = workbook.active
    db.connect()
    # Определение "рабочей зоны" (тут группа ячеек 4 на 3, в которой вся инфа для 1 пары 1 группы)
    work_zone = create_work_zone(sheet, 1)

    # Словарь для групп (нужен для того, чтобы знать в какой колонке, какая группа)
    groups_dict = {}
    print(f'{Fore.BLUE}Факультет: {faculty}{Style.RESET_ALL}')
    for num_week in range(1, 2 + 1):
        # print(num_week)
        # Цикл для прохода по одной группе
        for number_group in range(1, 10):
            # Название группы (Например прин-166)
            if groups_dict.get(number_group) is None:
                group_name = get_current_group_name(work_zone, sheet)
                groups_dict.update({number_group: group_name})
            else:
                group_name = groups_dict[number_group]
            if group_name == 'none':
                continue
            group_name_list = group_name.split(' ')
            # Если в ячейке написаны 2 группы, то прогоняем цикл по одному и тому же расписанию 2 раза, но с разными
            # названиями групп
            for group_index in range(len(group_name_list)):
                group_name = group_name_list[group_index]
                print(Fore.GREEN + f'Название группы: {group_name}' + Style.RESET_ALL)
                if course is None:
                    course = get_course(group_name)
                # Добавить в бд название группы
                insert_query = """
                    INSERT INTO groups (group_name, faculty, course, program)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (group_name) DO NOTHING
                """
                db.execute_query(insert_query, (group_name, faculty, course, study_program))

                # Цикл для прохода по всей неделе
                for week_day in range(1, 6 + 1):
                    print(f"День недели: {week_day}")
                    # Цикл для прохода по одному дню
                    for number_para in range(1, 6 + 1):
                        # print(f'Номер пары: {number_para}')
                        # Цикл для прохода по одной паре
                        new_has_lesson, new_lesson_dates, new_count_pars = None, None, None
                        has_lesson, lesson_dates, count_pars = None, None, None
                        for row in work_zone:
                            # Цикл для прохода по одной строке
                            for cell in row:
                                new_has_lesson, new_lesson_dates, new_count_pars = check_cell_has_data_and_dates(cell, sheet)
                                if new_has_lesson is not None:
                                    has_lesson = new_has_lesson
                                if new_lesson_dates is not None:
                                    lesson_dates = new_lesson_dates
                                if new_count_pars is not None:
                                    count_pars = new_count_pars
                        if count_pars is not None:
                            print(count_pars)


                        # Выбираем даты из таблицы dates по номеру недели и дню недели
                        query = "SELECT date FROM dates WHERE week_num = %s AND week_day = %s"
                        if lesson_dates is None:
                            dates = db.execute_query(query, (num_week, week_day))
                            dates = dates[1:]
                            dates = dates[0]
                            dates = [date[0] for date in dates]  # Преобразуем список кортежей в список значений
                        else:
                            dates = lesson_dates

                        # Для каждой выбранной даты вставляем записи в таблицу lessons
                        for date in dates:
                            insert_query = """
                                INSERT INTO lessons (group_name, lesson_order, is_busy, lesson_date)
                                SELECT %s, %s, %s, %s
                                WHERE NOT EXISTS (
                                    SELECT 1 FROM lessons 
                                    WHERE group_name = %s AND lesson_order = %s AND lesson_date = %s
                                )
                            """
                            values = (group_name, number_para, has_lesson, date, group_name, number_para, date)

                            db.execute_query(insert_query, values)
                        # print("_________________________")

                        work_zone = move_work_zone_down(work_zone, sheet)
                work_zone = move_work_zone_up(work_zone, sheet)
            work_zone = move_work_zone_right(work_zone, sheet)
        work_zone = create_work_zone(sheet, 7)
        # print(work_zone)
    db.disconnect()


def analyze_excel_files_in_folder(folder_path):
    exclude_prefix = '~$'  # Префикс для исключения файлов
    # Используем стек для хранения путей к папкам
    folders_path_stack = [folder_path]
    while folders_path_stack:
        current_folder_path = folders_path_stack.pop()
        if 'инженерных кадров' in current_folder_path:
            continue
        # print(Fore.CYAN, current_folder_path, Style.RESET_ALL)
        # Получаем список файлов и папок в текущей папке
        files = os.listdir(current_folder_path)
        for file in files:
            file_path = os.path.join(current_folder_path, file)
            if os.path.isdir(file_path):  # Если это папка
                folders_path_stack.append(file_path)
            elif file.endswith('.xlsx') and not file.startswith(exclude_prefix):
                # Обработка только файлов с расширением .xlsx и не начинающихся с указанного префикса
                analyze_excel_file(file_path, file)
    print(f"{Fore.GREEN}В базу занесены все группы!{Style.RESET_ALL}")


def delete_files_and_download_files():
    items = os.listdir()
    folders = [item for item in items if os.path.isdir(item) and "__" not in item]
    for folder in folders:
        try:
            shutil.rmtree(folder)
        except OSError as e:
            print(f'{Fore.RED}Ошибка при удалении папки {folder} {e} {Style.RESET_ALL}')
    download_schedule_files()
    convert_xls_to_xlsx()


if __name__ == '__main__':
    # Инициализация colorama (необходимо вызывать один раз в начале программы)
    init()

    db = PostgreSQLDatabase(host="localhost",
                            port="5433",
                            user='postgres',
                            password='postgres',
                            database="postgres")
    t = time.time()
    # delete_files_and_download_files()
    db.connect()
    db.truncate_table('lessons')
    db.truncate_table('groups')
    db.truncate_table('dates')
    db.disconnect()
    analyze_dates()
    analyze_excel_files_in_folder(
        'converted_files/')
    print("--- %s seconds ---" % (time.time() - t))
